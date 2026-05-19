import csv
import os
import time
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter


# cargar archivo csv o txt
def cargar_archivo(ruta):

  personas = []

  with open(ruta, newline="", encoding="utf-8") as archivo:

      lector = csv.DictReader(archivo)

      for fila in lector:

          try:
              persona = {
                  "id": int(fila["ID"]),
                  "nombre": fila["Nombre"] + " " + fila["Apellido"],
                  "edad": int(fila["Edad"]),
                  "puntaje_evaluacion": int(fila["Salario"])
              }

              personas.append(persona)

          except:
              print("error leyendo una fila")

  return personas


# cargar sqlite
def cargar_bd(ruta):

  datos = []

  conexion = sqlite3.connect(ruta)

  cursor = conexion.cursor()

  consulta = """
  SELECT id_persona,nombre,edad,puntaje_evaluacion
  FROM personas
  """

  cursor.execute(consulta)

  filas = cursor.fetchall()

  conexion.close()

  for fila in filas:

      persona = {
          "id": fila[0],
          "nombre": fila[1],
          "edad": fila[2],
          "puntaje_evaluacion": fila[3]
      }

      datos.append(persona)

  return datos


def cargar_datos(ruta):

  extension = os.path.splitext(ruta)[1].lower()

  if extension in [".csv", ".txt"]:
      return cargar_archivo(ruta)

  elif extension in [".db", ".sqlite", ".sqlite3"]:
      return cargar_bd(ruta)

  else:
      print("formato no valido")
      return []


# bubble sort
def bubble_sort(lista):

  datos = [x.copy() for x in lista]

  n = len(datos)

  for i in range(n):

      cambios = False

      for j in range(n - 1 - i):

          if datos[j]["puntaje_evaluacion"] > datos[j + 1]["puntaje_evaluacion"]:

              datos[j], datos[j + 1] = datos[j + 1], datos[j]

              cambios = True

      if cambios == False:
          break

  return datos


# insertion sort
def insertion_sort(lista):

  datos = [x.copy() for x in lista]

  for i in range(1, len(datos)):

      actual = datos[i]

      j = i - 1

      while j >= 0 and datos[j]["puntaje_evaluacion"] > actual["puntaje_evaluacion"]:

          datos[j + 1] = datos[j]

          j -= 1

      datos[j + 1] = actual

  return datos


# selection sort
def selection_sort(lista):

  datos = [x.copy() for x in lista]

  n = len(datos)

  for i in range(n - 1):

      menor = i

      for j in range(i + 1, n):

          if datos[j]["puntaje_evaluacion"] < datos[menor]["puntaje_evaluacion"]:

              menor = j

      if menor != i:

          aux = datos[i]
          datos[i] = datos[menor]
          datos[menor] = aux

  return datos


# complejidad
complejidades = {

  "Bubble Sort": {
      "peor": "O(n²)",
      "medio": "Θ(n²)",
      "mejor": "Ω(n)"
  },

  "Insertion Sort": {
      "peor": "O(n²)",
      "medio": "Θ(n²)",
      "mejor": "Ω(n)"
  },

  "Selection Sort": {
      "peor": "O(n²)",
      "medio": "Θ(n²)",
      "mejor": "Ω(n²)"
  }
}


def ejecutar(nombre, funcion, datos):

  print("\n----------------------------")
  print("algoritmo:", nombre)
  print("registros:", len(datos))
  print("----------------------------")

  inicio = time.perf_counter()

  resultado = funcion(datos)

  fin = time.perf_counter()

  tiempo_ms = (fin - inicio) * 1000

  print("\nordenamiento terminado")
  print(f"tiempo: {tiempo_ms:.4f} ms")

  info = complejidades[nombre]

  print("\ncomplejidad")
  print("peor caso:", info["peor"])
  print("caso medio:", info["medio"])
  print("mejor caso:", info["mejor"])

  return resultado, tiempo_ms


# guardar excel
def exportar_excel(datos, nombre_algoritmo, tiempo_ms, ruta_salida):

  libro = Workbook()

  hoja = libro.active

  hoja.title = "Resultados"

  encabezados = ["ID", "Nombre", "Edad", "Puntaje"]

  hoja.append(encabezados)

  fuente = Font(bold=True, color="FFFFFF")

  relleno = PatternFill("solid", start_color="4F81BD")

  borde = Border(
      left=Side(style="thin"),
      right=Side(style="thin"),
      top=Side(style="thin"),
      bottom=Side(style="thin")
  )

  for i in range(1, 5):

      celda = hoja.cell(row=1, column=i)

      celda.font = fuente
      celda.fill = relleno
      celda.border = borde
      celda.alignment = Alignment(horizontal="center")

  fila_excel = 2

  for persona in datos:

      hoja.append([
          persona["id"],
          persona["nombre"],
          persona["edad"],
          persona["puntaje_evaluacion"]
      ])

      for col in range(1, 5):

          hoja.cell(row=fila_excel, column=col).border = borde

      fila_excel += 1

  tamaños = [10, 30, 10, 20]

  for i, ancho in enumerate(tamaños, start=1):

      letra = get_column_letter(i)

      hoja.column_dimensions[letra].width = ancho

  reporte = libro.create_sheet("Reporte")

  reporte["A1"] = "Resumen"

  reporte["A2"] = "Algoritmo"
  reporte["B2"] = nombre_algoritmo

  reporte["A3"] = "Tiempo"
  reporte["B3"] = f"{tiempo_ms:.4f} ms"

  reporte["A4"] = "Total"
  reporte["B4"] = len(datos)

  reporte["A5"] = "Peor caso"
  reporte["B5"] = complejidades[nombre_algoritmo]["peor"]

  reporte["A6"] = "Caso medio"
  reporte["B6"] = complejidades[nombre_algoritmo]["medio"]

  reporte["A7"] = "Mejor caso"
  reporte["B7"] = complejidades[nombre_algoritmo]["mejor"]

  libro.save(ruta_salida)

  print("\nexcel guardado")


def menu():

  print("\n============================")
  print("ordenamiento ADA")
  print("============================")
  print("1. cargar datos")
  print("2. bubble sort")
  print("3. insertion sort")
  print("4. selection sort")
  print("5. salir")
  print("============================")


def main():

  datos = []

  while True:

      menu()

      opcion = input("\nopcion: ").strip()

      if opcion == "1":

          ruta = input("ruta archivo o bd: ").strip()

          if ruta == "":
              print("ruta vacia")
              continue

          if not os.path.exists(ruta):

              print("el archivo no existe")

              continue

          print("\ncargando datos...")

          datos = cargar_datos(ruta)

          print(f"datos cargados: {len(datos)}")

      elif opcion in ["2", "3", "4"]:

          if len(datos) == 0:

              print("primero cargue datos")

              continue

          algoritmos = {

              "2": ("Bubble Sort", bubble_sort),

              "3": ("Insertion Sort", insertion_sort),

              "4": ("Selection Sort", selection_sort)
          }

          nombre, funcion = algoritmos[opcion]

          resultado, tiempo = ejecutar(nombre, funcion, datos)

          archivo = nombre.lower().replace(" ", "_")

          exportar_excel(
              resultado,
              nombre,
              tiempo,
              f"resultado_{archivo}.xlsx"
          )

      elif opcion == "5":

          print("\nsaliendo...")

          break

      else:

          print("opcion invalida")


if __name__ == "__main__":
  main()